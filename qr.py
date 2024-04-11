import qrcode
from openpyxl import load_workbook
from pathlib import Path

# Function to clean and trim data
def clean_and_trim(data):
    if data is not None:
        return data.strip()
    return ""

# Function to generate and save QR codes
def generate_qr_codes(sheet, output_dir):
    # Ensure the output directory exists
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        email, name, srn, *_ = row
        
        # Skip empty or incomplete rows
        if not email or not name or not srn:
            continue
        
        # Clean and trim the name and SRN
        name = clean_and_trim(name)
        srn = clean_and_trim(srn).upper()  # Ensure SRN is in uppercase
        
        # Generate QR code
        qr = qrcode.make(srn)
        
        # Save the QR code image
        filename = f"{name}_{srn}.png"
        img_path = Path(output_dir) / filename
        qr.save(img_path)

# Specify the path to your Excel file and the output directory for QR codes
file_path = r"C:\Users\agarw\Downloads\TERRATHON 3.0 VOLS + CORE + MENTOR.xlsx"
output_directory = r"C:\Users\agarw\OneDrive\Desktop\volunters"

# Load the Excel workbook and select the sheet
workbook = load_workbook(filename=file_path)
sheet_name = workbook.sheetnames[0]  # Assuming the relevant data is in the first sheet
sheet = workbook[sheet_name]

# Generate QR codes for the volunteers
generate_qr_codes(sheet, output_directory)

print(f"QR codes have been generated and saved to {output_directory}")
